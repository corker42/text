# -*- coding: utf-8 -*-
"""
Created on Thu Aug  5 18:55:56 2021

@author: sanyuan
"""

import openpyxl
import requests
from lxml import etree


wb = openpyxl.Workbook() # 新建一个workbook
# print(str(wb))
ws = wb.active   # 调用正在运行的工作表
# print(ws)
ws['A1'] = '漫画';ws['B1'] = '作者名字';ws['C1'] = '人气'

number = 2
for n in range(293):
    url = f'https://ac.qq.com/Comic/all/search/hot/page/{n}'
    headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.111 Safari/537.36'
    }
    response = requests.get(url=url,headers=headers)
    response = response.text

    tree = etree.HTML(response)

    li_all = tree.xpath('/html/body/div[3]/div[2]/div/div[2]/ul/li')

    for q in range(len(li_all)):
        e = li_all[q].xpath('./div[2]/h3/a/text()')
        z = li_all[q].xpath('./div[2]/p[1]/text()')
        r = li_all[q].xpath('./div[2]/p[2]/span/em/text()')
        ws[f'A{number}'] = e[0]
        ws[f'B{number}'] = z[0]
        ws[f'C{number}'] = r[0]
        print('漫画:',e[0],'\n作者名字:',z[0],'\n人气:',r[0])
        print()
        number += 1
    wb.save('./漫画.xlsx')